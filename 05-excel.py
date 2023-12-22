import openpyxl

wb = openpyxl.load_workbook("planilla.xlsx")
hoja = wb.active
# print(hoja)
wb.create_sheet("Hoja3")
hoja3 = wb["Hoja3"]
hoja3.title = "Nuevo titulo"

# print(
#     hoja.max_row,
#     hoja.max_column
# )

celda = hoja["A1"]
celda.value = "Nombre Completo"
# print(celda.value)
celda2 = hoja.cell(row=2, column=1)
# print(
#     celda2.value,
#     celda2.row,
#     celda2.column,
#     celda2.coordinate
#     )

# for fila in range(1, hoja.max_row + 1):
#     for columna in range(1, hoja.max_column + 1):
#         celda = hoja.cell(row=fila, column=columna)
        # print(fila, columna, celda.value)

columna = hoja["A"]
row = hoja["1"]

# for col in columna:
#     print(col.value)

# print(columna)

hoja.append([1,2,3])
# for r in hoja.rows:
#     print(r)

# hoja.delete_rows(1, 1)

wb.save("nuevo_excel.xlsx")